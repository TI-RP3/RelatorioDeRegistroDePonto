from openpyxl import load_workbook

from Collaborator import Collaborator


class PontomaisReport:
    def __init__(self, table_path):
        self.table_path = table_path
        self.collaborators = []
        self.workDays = []

    def generateReport(table_path):
        workbook = load_workbook(table_path)
        worksheet = workbook.active
        
        for row in worksheet.iter_rows():
            for cell in row:
                    if (cell.value == "Colaborador"):
                        nomeColaborador = cell.offset(column=1).value
                        
                        colaborador = Collaborator(nomeColaborador,"","","","")  
                        print(colaborador.__str__())                      
