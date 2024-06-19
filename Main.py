from openpyxl import load_workbook
from Collaborator import Collaborator

def getDailyEntries(worksheet, period):
    general_list = []
    collaborators_infos = {}
    days_observed = []
    
    for i, row in enumerate(worksheet.iter_rows(min_row=3)):
        if row[0].value == "Colaborador":
            collab_name = row[1].value
            collaborators_infos[collab_name] = []
        elif (
            ("Seg" in row[0].value or
            "Ter" in row[0].value or
            "Qua" in row[0].value or
            "Qui" in row[0].value or
            "Sex" in row[0].value or
            "Sáb" in row[0].value or
            "Dom" in row[0].value) and
            (i > 1 and i <= period + 3)
        ):
            days_observed.append(row[0].value[5:10].replace("/", "."))                
        elif row[0].value not in ["Data", "Resumo", "TOTAIS"]:            
            first_entry = row[5].value
            first_exit = row[6].value
            second_entry = row[7].value
            second_exit = row[8].value
            collaborators_infos[collab_name].append(
                Collaborator(collab_name, first_entry, first_exit, second_entry, second_exit)
            )
    
    max_entries = max(len(entries) for entries in collaborators_infos.values())
    for _ in range(max_entries):
        general_list.append([])
        
    for i in range(len(days_observed)):
        general_list[i].append(days_observed[i])

    for _, entries in collaborators_infos.items():
        for j, entry in enumerate(entries):
            general_list[j].append(entry)

    return general_list

def readTemplateWorksheet(file_name):
    wb = load_workbook(file_name)
    template_ws = wb.active
    new_ws = wb.copy_worksheet(template_ws)
    new_ws.title = "01.05"
    wb.save(file_name)

def main():
    app_wb_name = "./Pontomais_-_Jornada_(01.05.2024_-_14.05.2024)_-_44082604.xlsx"
    template_wb_name = "./Relatório de Registro de ponto - Maio 2024.xlsx"
    #app_workbook = load_workbook(app_wb_name)
    #app_worksheet = app_workbook.active
    #period_observed = 14
    
    #daily_entries_per_collab = getDailyEntries(app_worksheet, period_observed)
    
    readTemplateWorksheet(template_wb_name)
    
    """ print("[\n")
    for _, daily_list in enumerate(daily_entries_per_collab):
        print(f"\t[ # dia: ")
        for collaborator in daily_list:
            print(collaborator.__str__() + "\n")
        print("\t],\n")
    print("]") """
    
    ...

if __name__ == "__main__":
    main()